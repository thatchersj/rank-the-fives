#!/usr/bin/env python3
"""
Create an Excel workbook with conditional formatting similar to the provided template.

Usage:
  python scripts/make_xlsx.py --csv docs/data/rankings_latest.csv --template data/template.xlsx --out docs/downloads/rankings_latest.xlsx
"""

from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def clear_sheet(ws, start_row: int = 1):
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", type=Path, required=True)
    ap.add_argument("--template", type=Path, required=True)
    ap.add_argument("--out", type=Path, required=True)
    args = ap.parse_args()

    # Load template to reuse its differential styles + conditional formatting colors
    wb = openpyxl.load_workbook(args.template)
    ws = wb.active

    # Read CSV
    import pandas as pd

    df = pd.read_csv(args.csv)

    # Build xlsx table columns: Rank, Initial, Surname, <tourney cols...>, then stats
    tourney_cols = [c for c in df.columns if len(c) == 4 and c[2] == " "]
    stat_cols = [c for c in df.columns if c not in (["Initial", "Surname"] + tourney_cols)]
    # Expected stat order (same as template)
    stat_order = ["RPA", "POSS", "Played", "MissedLast", "PC", "PC2", "RANK", "RANK2", "RANK3"]
    stat_cols = [c for c in stat_order if c in stat_cols]

    out_df = df.copy()
    out_df.insert(0, "Rank", out_df["RANK3"].astype(int))
    out_df = out_df[["Rank", "Initial", "Surname"] + tourney_cols + stat_cols]

    # Ensure sheet has enough columns: if df has more than template, extend header and widths (simple)
    needed_cols = out_df.shape[1]
    if ws.max_column < needed_cols:
        # Add extra columns at the end
        for c in range(ws.max_column + 1, needed_cols + 1):
            ws.cell(row=1, column=c).value = None

    # Clear existing contents
    clear_sheet(ws, start_row=1)

    # Write header
    for c, col_name in enumerate(out_df.columns, start=1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.font = Font(bold=True)

    # Write data
    for r_idx, row in enumerate(out_df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Column widths: reuse template defaults for first 48 columns; set a reasonable width for extras
    for c in range(1, needed_cols + 1):
        letter = get_column_letter(c)
        if letter not in ws.column_dimensions or ws.column_dimensions[letter].width is None:
            ws.column_dimensions[letter].width = 6

    # Rebuild conditional formatting for results area (Round cells)
    # We'll recreate rules (same values) and rely on template's existing differential styles.
    ws.conditional_formatting._cf_rules.clear()

    # Determine range for round cells: columns D.. up to last tourney col
    first_round_col = 4  # D
    last_round_col = 3 + len(tourney_cols)
    last_row = out_df.shape[0] + 1  # header row + data rows

    if last_round_col >= first_round_col and last_row >= 2:
        start = f"{get_column_letter(first_round_col)}2"
        end = f"{get_column_letter(last_round_col)}{last_row}"
        rng = f"{start}:{end}"

        # These dxfIds correspond to the template's differential styles:
        # P -> 0, W -> 1, F -> 2, SF -> 3, QF -> 4, L16 -> 5, NQ -> 6
        mapping = [("NQ", 6, 2), ("L16", 5, 3), ("QF", 4, 4), ("SF", 3, 5), ("F", 2, 6), ("W", 1, 7), ("P", 0, 8)]
        for val, dxf_id, priority in mapping:
            rule = Rule(type="cellIs", operator="equal", dxfId=dxf_id, priority=priority, formula=[f'"{val}"'])
            ws.conditional_formatting.add(rng, rule)

    # Save
    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
